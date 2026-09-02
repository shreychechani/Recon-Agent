"""Read the three messy input files into canonical Pydantic models.

Handles: junk header rows (xlsx), non-obvious column names (csv), nested JSON with
comma-separated string amounts, and three different date formats.

RUPEES -> PAISE CONVERSION HAPPENS HERE AND ONLY HERE. Amounts are parsed as
``Decimal`` and multiplied by 100 to integer paise. There is no ``float`` in this
path — floats lose paise.
"""

from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from src.ingest import schema_map
from src.models import BankCredit, Order, SettlementLine

# --------------------------------------------------------------------------- #
# Scalar parsers
# --------------------------------------------------------------------------- #

_DATE_FORMATS = [
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d",
    "%d/%m/%Y",
    "%d-%m-%Y",
    "%d-%b-%y",
    "%d-%B-%Y",
    "%d %b %Y",
    "%Y/%m/%d",
]


def parse_money_to_paise(raw: object) -> int:
    """'49,847.00' / '-498.47' / '₹1,200' -> integer paise, via Decimal."""
    if raw is None:
        return 0
    s = str(raw).strip().replace(",", "").replace("₹", "").replace("INR", "").strip()
    if s == "" or s.lower() == "none":
        return 0
    return int((Decimal(s) * 100).to_integral_value())


def parse_dt(raw: object) -> datetime:
    s = str(raw).strip()
    for fmt in _DATE_FORMATS:
        for candidate in (s, s.title()):  # .title() rescues 'jan' vs 'Jan'
            try:
                return datetime.strptime(candidate, fmt)
            except ValueError:
                continue
    raise ValueError(f"unrecognized date/datetime: {s!r}")


def parse_date(raw: object) -> date:
    return parse_dt(raw).date()


# --------------------------------------------------------------------------- #
# Generic readers -> (headers, list[dict keyed by source header])
# --------------------------------------------------------------------------- #


def _find_header_row(rows: list[list[str]], file_type: str) -> int:
    """Locate the real header row by counting alias-matchable cells."""
    best_idx, best_score = 0, -1
    for i, row in enumerate(rows[:10]):  # header is always near the top
        norm_cells = {schema_map._norm(c) for c in row if c}
        score = 0
        for aliases in schema_map.ALIASES[file_type].values():
            if any(schema_map._norm(a) in norm_cells for a in aliases):
                score += 1
        if score > best_score:
            best_idx, best_score = i, score
    return best_idx


def _read_xlsx(path: Path, file_type: str) -> tuple[list[str], list[dict]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = [["" if c is None else str(c) for c in r] for r in ws.iter_rows(values_only=True)]
    wb.close()
    h = _find_header_row(rows, file_type)
    headers = [c for c in rows[h]]
    data = [
        dict(zip(headers, r))
        for r in rows[h + 1 :]
        if any(str(c).strip() for c in r)
    ]
    return headers, data


def _read_csv(path: Path) -> tuple[list[str], list[dict]]:
    import csv

    with path.open(newline="") as f:
        reader = csv.DictReader(f)
        headers = list(reader.fieldnames or [])
        data = [dict(row) for row in reader]
    return headers, data


def _read_bank_json(path: Path) -> tuple[list[str], list[dict]]:
    payload = json.loads(path.read_text())
    txns = payload["transactions"] if isinstance(payload, dict) else payload
    headers = list(txns[0].keys()) if txns else []
    return headers, txns


# --------------------------------------------------------------------------- #
# Canonical loaders
# --------------------------------------------------------------------------- #


def _get(row: dict, mapping: dict, field: str):
    col = mapping.get(field)
    return row.get(col) if col else None


def load_orders(path: str | Path) -> list[Order]:
    path = Path(path)
    headers, rows = _read_xlsx(path, "orders")
    m = schema_map.map_columns("orders", headers, rows)
    out: list[Order] = []
    for r in rows:
        status = str(_get(r, m, "status") or "captured")
        if status not in ("captured", "refunded", "partially_refunded", "failed"):
            status = "captured"
        out.append(
            Order(
                order_id=str(_get(r, m, "order_id")),
                amount_paise=parse_money_to_paise(_get(r, m, "amount")),
                created_at=parse_dt(_get(r, m, "created_at")),
                status=status,  # type: ignore[arg-type]
                customer_ref=str(_get(r, m, "customer_ref") or ""),
            )
        )
    return out


def load_settlements(path: str | Path) -> list[SettlementLine]:
    path = Path(path)
    headers, rows = _read_csv(path)
    m = schema_map.map_columns("settlements", headers, rows)
    out: list[SettlementLine] = []
    for r in rows:
        gross = parse_money_to_paise(_get(r, m, "gross"))
        net_raw = _get(r, m, "net")
        net = parse_money_to_paise(net_raw)
        order_id = _get(r, m, "order_id")
        line_type = str(_get(r, m, "line_type") or "sale")
        if line_type not in ("sale", "refund", "chargeback", "adjustment"):
            line_type = "sale"
        out.append(
            SettlementLine(
                settlement_id=str(_get(r, m, "settlement_id")),
                payout_batch_id=str(_get(r, m, "payout_batch_id")),
                order_id=str(order_id) if order_id else None,
                gross_paise=gross,
                fee_paise=parse_money_to_paise(_get(r, m, "fee")),
                gst_paise=parse_money_to_paise(_get(r, m, "gst")),
                net_paise=net,
                line_type=line_type,  # type: ignore[arg-type]
                settled_at=parse_dt(_get(r, m, "settled_at")),
                payout_ref=(str(_get(r, m, "payout_ref")) or None) if _get(r, m, "payout_ref") else None,
            )
        )
    return out


def load_bank(path: str | Path) -> list[BankCredit]:
    path = Path(path)
    headers, rows = _read_bank_json(path)
    m = schema_map.map_columns("bank", headers, rows)
    out: list[BankCredit] = []
    for r in rows:
        utr = _get(r, m, "utr")
        out.append(
            BankCredit(
                bank_txn_id=str(_get(r, m, "bank_txn_id")),
                amount_paise=parse_money_to_paise(_get(r, m, "amount")),
                value_date=parse_date(_get(r, m, "value_date")),
                narration=str(_get(r, m, "narration") or ""),
                utr=str(utr) if utr else None,
            )
        )
    return out


def load_dataset(data_dir: str | Path) -> tuple[list[Order], list[SettlementLine], list[BankCredit]]:
    d = Path(data_dir)
    return (
        load_orders(d / "orders.xlsx"),
        load_settlements(d / "settlements.csv"),
        load_bank(d / "bank.json"),
    )
